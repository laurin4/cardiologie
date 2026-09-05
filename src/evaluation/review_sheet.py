"""Build a minimal review sheet for manual consistency checks (CSV or Excel)."""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Sequence

CLINICAL_COLUMNS = [
    "pacemaker",
    "atrial_fibrillation",
    "cerebrovascular_event",
    "reoperation_required",
    "reoperation_context",
    "multi_system_failure",
    "rethoracotomy",
    "rethoracotomy_context",
    "liver_cirrhosis",
]

COLUMNS = [
    "patient_id",
    "fall_nummers",
    "verlegung_fallnr",
    "verlegung_matched",
    "status",
    *CLINICAL_COLUMNS,
    "information_sufficient",
    "evidence_quotes",
    "reasoning",
    "correct_reop",
    "notes",
]

LOGGER = logging.getLogger(__name__)


def _clean(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _quotes_cell(raw: object) -> str:
    s = _clean(raw)
    if not s:
        return ""
    try:
        parsed = json.loads(s)
        if isinstance(parsed, list):
            return " | ".join(_clean(x) for x in parsed if _clean(x))
    except Exception:
        pass
    return s.replace("\n", " | ")


def load_result_rows(csv_path: Path) -> list[dict]:
    with open(csv_path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def enrich_fall_keys_from_raw(
    result_rows: list[dict],
    *,
    diagnose_paths: Optional[Sequence[Path]] = None,
    verlegung_paths: Optional[Sequence[Path]] = None,
) -> list[dict]:
    """
    Fill ``fall_nummers`` / ``verlegung_fallnr`` / ``verlegung_matched`` on
    existing result rows by looking up HER raw tables (no LLM re-run).
    """
    from src.preprocessing.diagnose_loader import load_patient_diagnoseliste_many
    from src.preprocessing.report_identity import normalize_str
    from src.preprocessing.report_loader import (
        discover_her_diagnose_paths,
    )
    from src.preprocessing.verlegung_loader import (
        discover_her_verlegung_paths,
        load_verlegung_by_fall,
        pick_verlegung_for_falls,
    )

    diag_paths = list(diagnose_paths) if diagnose_paths is not None else discover_her_diagnose_paths()
    verl_paths = (
        list(verlegung_paths) if verlegung_paths is not None else discover_her_verlegung_paths()
    )
    if not diag_paths:
        LOGGER.warning("No HER_Diagnose* under data/raw/; cannot enrich FallNummer.")
        return result_rows

    by_patient: Dict[str, List[str]] = {}
    for rec in load_patient_diagnoseliste_many(diag_paths):
        pid = normalize_str(rec.get("patient_id") or rec.get("report_id") or "")
        if not pid:
            continue
        by_patient[pid] = list(rec.get("fall_nummers") or [])

    by_fall = load_verlegung_by_fall(verl_paths) if verl_paths else {}

    n_filled = 0
    n_matched = 0
    sample_miss: List[str] = []
    for row in result_rows:
        pid = normalize_str(row.get("patient_id") or row.get("report_id") or "")
        falls = by_patient.get(pid) or []
        if not falls and pid:
            # Retry after Excel-style normalization already applied; try int-strip.
            if pid.isdigit():
                falls = by_patient.get(str(int(pid))) or []
        if falls:
            row["fall_nummers"] = " | ".join(falls)
            n_filled += 1
        elif not _clean(row.get("fall_nummers")):
            row["fall_nummers"] = ""

        info = pick_verlegung_for_falls(by_fall, falls) if falls else None
        if info:
            row["verlegung_fallnr"] = str(info.get("verlegung_fallnr") or "")
            row["verlegung_matched"] = "True"
            n_matched += 1
        else:
            if not _clean(row.get("verlegung_fallnr")):
                row["verlegung_fallnr"] = ""
            row["verlegung_matched"] = "False"
            if len(sample_miss) < 5:
                sample_miss.append(f"{pid or '?'} falls={falls or '-'}")

    msg = (
        f"Enriched FallNummer for {n_filled}/{len(result_rows)} result rows "
        f"({n_matched} Verlegung matches) from {len(diag_paths)} Diagnose file(s), "
        f"{len(by_fall)} Verlegung FallNummer(n)"
    )
    LOGGER.info(msg)
    print(msg)
    if n_matched == 0 and result_rows:
        print(
            "NOTE: 0 Verlegung matches among these result patients. "
            "Your run likely took the first N Diagnose patients, who often have "
            "no overlapping FallNummer. Re-run with --require-verlegung, or "
            "spot-check Diagnoseliste-only until then."
        )
        if sample_miss:
            print("Sample unmatched:", "; ".join(sample_miss))
    return result_rows


def build_review_rows(result_rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    for row in result_rows:
        patient_id = _clean(row.get("patient_id")) or _clean(row.get("report_id"))
        if not patient_id:
            patient_id = _clean(row.get("source_row_id")) or _clean(row.get("report_name"))
        item = {
            "patient_id": patient_id,
            "fall_nummers": _clean(row.get("fall_nummers")),
            "verlegung_fallnr": _clean(row.get("verlegung_fallnr")),
            "verlegung_matched": _clean(row.get("verlegung_matched")),
            "status": _clean(row.get("status")),
            "information_sufficient": _clean(row.get("information_sufficient")),
            "evidence_quotes": _quotes_cell(row.get("evidence_quotes")),
            "reasoning": _clean(row.get("reasoning")).replace("\n", " "),
            "correct_reop": "",
            "notes": "",
        }
        for col in CLINICAL_COLUMNS:
            item[col] = _clean(row.get(col))
        out.append(item)
    return out


def write_csv(rows: list[dict], out_path: Path) -> None:
    """Write semicolon-separated CSV (Excel-friendly on DE Windows)."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=COLUMNS,
            delimiter=";",
            extrasaction="ignore",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        writer.writerows(rows)


def write_excel(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(name, ""))
            cell.alignment = wrap

    widths = {
        "patient_id": 14,
        "fall_nummers": 18,
        "verlegung_fallnr": 16,
        "verlegung_matched": 12,
        "status": 12,
        "evidence_quotes": 40,
        "reasoning": 40,
        "reoperation_context": 40,
        "rethoracotomy_context": 40,
        "notes": 30,
    }
    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
